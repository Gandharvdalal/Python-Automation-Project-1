import openpyxl
workbook=openpyxl.load_workbook(filename='inventory.xlsx')
worksheet=workbook['Sheet1']
products_per_supplier={}
total_value_per_supplier={}

for i in range(2,worksheet.max_row+1):
    suppliername=worksheet.cell(i,4).value
    inventory_price=worksheet.cell(i,5)
    if suppliername in products_per_supplier:
        temp=products_per_supplier[suppliername]
        products_per_supplier[suppliername]=temp+1
    else:
        products_per_supplier[suppliername]=1


    # for finding products_per_supplier
    price=worksheet.cell(i,3).value
    inventory=worksheet.cell(i,2).value
    if suppliername in total_value_per_supplier:
        temp2=total_value_per_supplier[suppliername]
        total_value_per_supplier[suppliername]=temp2+inventory*price
    else:
        total_value_per_supplier[suppliername]=inventory*price  

    # add value for total inventory price
    inventory_price.value=inventory*price
    # and save the changes
    #workbook.save(filename='inventory.xlsx')
    workbook.save("inventory_with_total_value.xlsx") #it saves the file store in the new file


print(products_per_supplier)
print(total_value_per_supplier)